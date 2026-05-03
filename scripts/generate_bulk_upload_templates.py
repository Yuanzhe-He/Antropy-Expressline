from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "templates" / "bulk-upload" / "express-line-bulk-upload-template.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="111111")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="F3F3F0")
THIN_BORDER = Border(bottom=Side(style="thin", color="C9C9C4"))

LIST_VALUES = {
    "currency": ["MXN", "USD"],
    "boolean": ["TRUE", "FALSE"],
    "module": ["handover", "customs", "inland"],
    "price_mode": ["pretax", "aftertax"],
    "cutoff_handler": [
        "customs_broker_only",
        "customer_only",
        "broker_or_consignee",
    ],
    "charge_scope": ["bl", "container"],
    "yard_fee_type": ["dropoff", "customs"],
}

SHEETS = [
    {
        "name": "shipping_lines",
        "columns": [
            ("shipping_line_id", "Stable unique ID. Use lowercase letters, numbers, and hyphens."),
            ("name", "Shipping line display name."),
            ("enabled", "TRUE or FALSE."),
            ("invoice_to_consignee_only", "TRUE if invoices can only be issued to consignee/customer."),
            ("invoice_note", "Optional note shown to sales."),
            ("demurrage_cutoff_handler", "customs_broker_only, customer_only, or broker_or_consignee."),
        ],
        "validations": {
            "enabled": "boolean",
            "invoice_to_consignee_only": "boolean",
            "demurrage_cutoff_handler": "cutoff_handler",
        },
    },
    {
        "name": "container_types",
        "columns": [
            ("container_type_key", "Stable container type key, e.g. 20gp or 40hq."),
            ("label", "User-facing container type label."),
            ("module", "handover, customs, inland, or leave blank for shared usage."),
            ("enabled", "TRUE or FALSE."),
        ],
        "validations": {"module": "module", "enabled": "boolean"},
    },
    {
        "name": "tax_presets",
        "columns": [
            ("module", "Module where this tax preset applies."),
            ("tax_id", "Stable tax preset ID, e.g. vat-16."),
            ("label", "Display label, e.g. 16%."),
            ("rate", "Decimal rate. Use 0.16 for 16%."),
        ],
        "validations": {"module": "module"},
    },
    {
        "name": "handover_local_charges",
        "columns": [
            ("shipping_line_id", "Must exist in shipping_lines."),
            ("charge_id", "Stable charge ID."),
            ("concept", "Charge name shown in results."),
            ("charge_scope", "bl or container."),
            ("container_type_key", "Required when charge_scope is container."),
            ("amount", "Pretax amount."),
            ("currency", "MXN or USD."),
            ("tax_rate", "Decimal rate. Use 0.16 for 16%."),
            ("note", "Optional note."),
        ],
        "validations": {"charge_scope": "charge_scope", "currency": "currency"},
    },
    {
        "name": "handover_guarantees",
        "columns": [
            ("shipping_line_id", "Must exist in shipping_lines."),
            ("benefit_enabled", "TRUE when guarantee/deposit is waived."),
            ("benefit_expires_at", "Optional ISO date YYYY-MM-DD."),
            ("container_type_key", "Container type for deposit rate."),
            ("amount", "Pretax amount. Use 0 when waived by rule."),
            ("currency", "MXN or USD."),
            ("tax_rate", "Decimal rate."),
            ("note", "Optional note."),
        ],
        "validations": {"benefit_enabled": "boolean", "currency": "currency"},
    },
    {
        "name": "handover_demurrage",
        "columns": [
            ("shipping_line_id", "Must exist in shipping_lines."),
            ("container_type_key", "Container type key."),
            ("start_day", "Start day. Importer should validate sequence."),
            ("end_day", "End day. Leave blank only for final open-ended tier."),
            ("amount_per_day", "Daily pretax amount. Use 0 for free days."),
            ("currency", "MXN or USD."),
            ("tax_rate", "Decimal rate."),
            ("note", "Optional note."),
        ],
        "validations": {"currency": "currency"},
    },
    {
        "name": "customs_ports_terminals",
        "columns": [
            ("port_id", "Stable port ID."),
            ("port_name", "Port display name."),
            ("terminal_id", "Stable terminal ID."),
            ("terminal_name", "Terminal display name."),
            ("enabled", "TRUE or FALSE."),
            ("note", "Optional note."),
        ],
        "validations": {"enabled": "boolean"},
    },
    {
        "name": "customs_yards",
        "columns": [
            ("yard_id", "Stable yard ID."),
            ("yard_name", "Yard display name."),
            ("port_id", "Allowed port ID. One row per yard-port mapping."),
            ("shipping_line_id", "Allowed shipping line ID. One row per yard-shipping line mapping."),
            ("enabled", "TRUE or FALSE."),
            ("note", "Optional note."),
        ],
        "validations": {"enabled": "boolean"},
    },
    {
        "name": "customs_terminal_fees",
        "columns": [
            ("port_id", "Must exist in customs_ports_terminals."),
            ("terminal_id", "Must exist in customs_ports_terminals."),
            ("fee_id", "Stable fee ID."),
            ("concept", "Fee name shown in results."),
            ("container_type_key", "Container type key."),
            ("amount", "Pretax amount per container."),
            ("currency", "MXN or USD."),
            ("tax_rate", "Decimal rate."),
            ("note", "Optional note."),
        ],
        "validations": {"currency": "currency"},
    },
    {
        "name": "customs_storage_rules",
        "columns": [
            ("port_id", "Must exist in customs_ports_terminals."),
            ("terminal_id", "Must exist in customs_ports_terminals."),
            ("container_type_key", "Container type key."),
            ("start_day", "Start day. Importer should validate sequence."),
            ("end_day", "End day. Leave blank only for final open-ended tier."),
            ("amount_per_day", "Daily pretax amount. Use 0 for free days."),
            ("currency", "MXN or USD."),
            ("tax_rate", "Decimal rate."),
            ("note", "Optional note."),
        ],
        "validations": {"currency": "currency"},
    },
    {
        "name": "customs_yard_fees",
        "columns": [
            ("yard_id", "Must exist in customs_yards."),
            ("fee_type", "dropoff or customs."),
            ("fee_id", "Stable fee ID."),
            ("concept", "Fee name shown in results."),
            ("container_type_key", "Container type key."),
            ("amount", "Pretax amount per container."),
            ("currency", "MXN or USD."),
            ("tax_rate", "Decimal rate."),
            ("note", "Optional note."),
        ],
        "validations": {"fee_type": "yard_fee_type", "currency": "currency"},
    },
]


def apply_header_style(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


def add_list_sheet(workbook: Workbook) -> dict[str, str]:
    sheet = workbook.create_sheet("lists")
    ranges: dict[str, str] = {}

    for col_index, (list_name, values) in enumerate(LIST_VALUES.items(), start=1):
        sheet.cell(row=1, column=col_index, value=list_name)
        apply_header_style(sheet.cell(row=1, column=col_index))
        for row_index, value in enumerate(values, start=2):
            sheet.cell(row=row_index, column=col_index, value=value)
        column_letter = sheet.cell(row=1, column=col_index).column_letter
        ranges[list_name] = f"lists!${column_letter}$2:${column_letter}${len(values) + 1}"
        sheet.column_dimensions[column_letter].width = max(
            18,
            max(len(str(value)) for value in values) + 2,
        )

    sheet.sheet_state = "hidden"
    return ranges


def add_readme_sheet(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "README"
    rows = [
        ["Purpose", "This workbook is a bulk upload template only. It is not the system source of truth."],
        ["Source of truth", "The application data store remains the app database/data files. Imports should validate and preview changes before writing."],
        ["IDs", "Use stable IDs/keys. Do not rename IDs casually; names/labels can change."],
        ["Currencies", "Use MXN or USD."],
        ["Taxes", "Use decimal rates. 16% should be entered as 0.16."],
        ["Free days", "Use amount_per_day = 0 for free tiers."],
        ["Open-ended tiers", "Leave end_day blank only on the final tier."],
        ["Recommended import flow", "Upload -> validate -> preview diff -> confirm -> write -> audit log."],
    ]

    sheet.append(["Field", "Instruction"])
    for cell in sheet[1]:
        apply_header_style(cell)
    for row in rows:
        sheet.append(row)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 110
    for row in sheet.iter_rows(min_row=2):
        row[0].fill = SUBTLE_FILL
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_data_sheet(workbook: Workbook, definition: dict, list_ranges: dict[str, str]) -> None:
    sheet = workbook.create_sheet(definition["name"])
    sheet.freeze_panes = "A2"

    for col_index, (header, comment) in enumerate(definition["columns"], start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        apply_header_style(cell)
        cell.comment = Comment(comment, "Antropy AI")
        sheet.column_dimensions[cell.column_letter].width = min(max(len(header) + 4, 16), 36)

    for header, list_name in definition.get("validations", {}).items():
        col_index = [item[0] for item in definition["columns"]].index(header) + 1
        col_letter = sheet.cell(row=1, column=col_index).column_letter
        validation = DataValidation(
            type="list",
            formula1=f"={list_ranges[list_name]}",
            allow_blank=True,
        )
        sheet.add_data_validation(validation)
        validation.add(f"{col_letter}2:{col_letter}1000")

    for row in range(2, 1001):
        for col in range(1, len(definition["columns"]) + 1):
            sheet.cell(row=row, column=col).border = THIN_BORDER


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    add_readme_sheet(workbook)
    list_ranges = add_list_sheet(workbook)

    for definition in SHEETS:
        add_data_sheet(workbook, definition, list_ranges)

    workbook.save(OUTPUT)
    print(f"bulk-upload-template-created: {OUTPUT}")


if __name__ == "__main__":
    main()
