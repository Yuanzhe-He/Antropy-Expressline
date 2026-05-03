from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(
    "/Users/yuanzhehe/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/wxid_8b9mu5g3fomo22_21e8/temp/drag/TARIFARIO 120426.xlsx"
)
OUTPUT = ROOT / "data" / "shipping-lines.json"

SKIP_HEADERS = {
    None,
    "",
    "USD",
    "MXN",
    "Subtotal",
    "TOTAL",
    "Calculo",
    "IVA",
    "Nota",
    "MANZANILLO",
    "Facturación",
}


@dataclass
class GroupInfo:
    key: str
    label: str
    col: int
    currency_cols: tuple[int, int]


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "item"


def normalize_currency(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_tax(value: Any) -> float:
    if value is None:
        return 1.0
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text or text.upper() == "EX":
        return 1.0
    try:
        return float(text)
    except ValueError:
        return 1.0


def tax_rate_from_multiplier(value: float) -> float:
    return round(max(0, value - 1.0), 4) if value > 1 else 0.0


def extract_groups(ws) -> tuple[list[GroupInfo], int | None]:
    groups: list[GroupInfo] = []
    bl_col = None
    row = 2

    for col in range(1, ws.max_column + 1):
        label = clean_text(ws.cell(row=row, column=col).value)
        if label == "BL":
            bl_col = col
            break
        if label in SKIP_HEADERS:
            continue
        next_label = clean_text(ws.cell(row=row, column=col + 1).value)
        next_next_label = clean_text(ws.cell(row=row, column=col + 2).value)
        if next_label == "USD" or next_next_label == "USD":
            groups.append(
                GroupInfo(
                    key=slugify(label),
                    label=label,
                    col=col,
                    currency_cols=(col + 1, col + 2),
                )
            )

    if groups:
        return groups, bl_col

    current_parent = None
    for col in range(1, ws.max_column + 1):
        parent_label = clean_text(ws.cell(row=2, column=col).value)
        child_label = clean_text(ws.cell(row=3, column=col).value)
        if parent_label == "BL" or child_label == "BL":
            bl_col = col
            break
        if parent_label not in SKIP_HEADERS:
            current_parent = parent_label
        if child_label in SKIP_HEADERS:
            continue
        next_label = clean_text(ws.cell(row=3, column=col + 1).value)
        next_next_label = clean_text(ws.cell(row=3, column=col + 2).value)
        if next_label == "USD" or next_next_label == "USD":
            label = f"{current_parent} {child_label}".strip() if current_parent else child_label
            groups.append(
                GroupInfo(
                    key=slugify(label),
                    label=label,
                    col=col,
                    currency_cols=(col + 1, col + 2),
                )
            )

    return groups, bl_col


def extract_group_rates(ws, row_idx: int, groups: list[GroupInfo]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for group in groups:
        usd = normalize_currency(ws.cell(row=row_idx, column=group.currency_cols[0]).value)
        mxn = normalize_currency(ws.cell(row=row_idx, column=group.currency_cols[1]).value)
        qty_hint = normalize_currency(ws.cell(row=row_idx, column=group.col).value)
        if usd is None and mxn is None:
            continue
        currency = "USD" if usd is not None else "MXN"
        result[group.key] = {
            "label": group.label,
            "qtyHint": qty_hint or 1,
            "currency": currency,
            "rate": usd if usd is not None else mxn,
        }
    return result


def extract_bl_rate(ws, row_idx: int, bl_col: int | None) -> dict[str, Any] | None:
    if bl_col is None:
        return None
    usd = normalize_currency(ws.cell(row=row_idx, column=bl_col + 1).value)
    mxn = normalize_currency(ws.cell(row=row_idx, column=bl_col + 2).value)
    qty_hint = normalize_currency(ws.cell(row=row_idx, column=bl_col).value)
    if usd is None and mxn is None:
        return None
    return {
        "qtyHint": qty_hint or 1,
        "currency": "USD" if usd is not None else "MXN",
        "rate": usd if usd is not None else mxn,
    }


def extract_terminal_mix(ws) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for col in range(1, ws.max_column):
        label = clean_text(ws.cell(row=2, column=col).value)
        ratio = normalize_currency(ws.cell(row=2, column=col + 1).value)
        if not label or ratio is None:
            continue
        if label in {"MANZANILLO", "Facturación", "IVA", "Nota", "BL", "TOTAL", "Calculo"}:
            continue
        if 0 <= ratio <= 1:
            entries.append({"terminal": label, "ratio": ratio})
    return entries


def extract_invoice_note(ws) -> tuple[bool, str | None]:
    for row_idx in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column):
            label = clean_text(ws.cell(row=row_idx, column=col).value)
            note = clean_text(ws.cell(row=row_idx, column=col + 1).value)
            if label != "Facturación" or not note:
                continue
            only_consignee = "consignad" in note.lower() or "consignee" in note.lower()
            return only_consignee, note
    return False, None


def extract_free_days(ws, groups: list[GroupInfo]) -> dict[str, Any]:
    result = {"defaultDays": 0, "daysByGroup": {}}
    for row_idx in range(1, ws.max_row + 1):
        marker = clean_text(ws.cell(row=row_idx, column=3).value)
        if marker != "Corte de demoras":
            continue
        for group in groups:
            days = normalize_currency(ws.cell(row=row_idx, column=group.col).value)
            if days is not None:
                result["daysByGroup"][group.key] = int(days)
        if result["daysByGroup"]:
            result["defaultDays"] = int(next(iter(result["daysByGroup"].values())))
        return result
    return result


def extract_lines(ws, title: str) -> dict[str, Any]:
    groups, bl_col = extract_groups(ws)
    only_consignee, invoice_note = extract_invoice_note(ws)
    terminal_mix = extract_terminal_mix(ws)

    local_charges = []
    guarantee = {
        "benefitEnabled": False,
        "benefitExpiresAt": None,
        "benefitNote": None,
        "taxRate": 0,
        "ratesByGroup": {},
        "fallbackRatesByGroup": {},
        "blRate": None,
    }
    demurrage_tiers = []
    demurrage_free_days = extract_free_days(ws, groups)

    section = "localCharges"
    for row_idx in range(3, ws.max_row + 1):
        section_title = clean_text(ws.cell(row=row_idx, column=2).value)
        concept = clean_text(ws.cell(row=row_idx, column=3).value)
        tax_multiplier = normalize_tax(ws.cell(row=row_idx, column=4).value)
        note = clean_text(ws.cell(row=row_idx, column=5).value)
        group_rates = extract_group_rates(ws, row_idx, groups)
        bl_rate = extract_bl_rate(ws, row_idx, bl_col)

        if section_title and section_title.startswith("Garantia"):
            section = "guarantee"
            guarantee["benefitNote"] = note
            guarantee["benefitEnabled"] = bool(note and "beneficio" in note.lower() and "no hay" not in note.lower())
            guarantee["taxRate"] = tax_rate_from_multiplier(tax_multiplier)
            guarantee["ratesByGroup"] = {
                key: {**value, "taxMultiplier": tax_multiplier}
                for key, value in group_rates.items()
            }
            guarantee["fallbackRatesByGroup"] = dict(guarantee["ratesByGroup"])
            guarantee["blRate"] = {**bl_rate, "taxMultiplier": tax_multiplier} if bl_rate else None
            continue

        if section_title == "Demoras":
            section = "demurrage"

        if section == "localCharges":
            if not concept and not bl_rate and not group_rates:
                continue
            local_charges.append(
                {
                    "id": f"{slugify(title)}-{row_idx}",
                    "concept": concept or f"Cargo {row_idx}",
                    "note": note,
                    "taxRate": tax_rate_from_multiplier(tax_multiplier),
                    "taxMultiplier": tax_multiplier,
                    "groupRates": group_rates,
                    "blRate": bl_rate,
                }
            )
            continue

        if section == "demurrage":
            if concept == "Corte de demoras":
                break
            if concept and concept.startswith("Los días libres"):
                break
            if not note and not group_rates:
                continue
            if note == "Días libres":
                continue
            demurrage_tiers.append(
                {
                    "id": f"{slugify(title)}-demurrage-{row_idx}",
                    "label": note or f"Tier {row_idx}",
                    "note": concept,
                    "taxRate": tax_rate_from_multiplier(tax_multiplier),
                    "groupRates": {
                        key: {**value, "taxMultiplier": tax_multiplier}
                        for key, value in group_rates.items()
                    },
                }
            )

    container_groups = [
        {"key": group.key, "label": group.label}
        for group in groups
    ]

    return {
        "id": slugify(title),
        "name": title,
        "active": True,
        "containerGroups": container_groups,
        "invoiceToConsigneeOnly": only_consignee,
        "invoiceNote": invoice_note,
        "demurrageCutoffHandledBy": "customs_broker_only",
        "terminalMix": terminal_mix,
        "localCharges": local_charges,
        "guarantee": guarantee,
        "demurrage": {
            "freeDays": demurrage_free_days,
            "tiers": demurrage_tiers,
        },
        "notes": {
            "sourceSheet": title,
            "code": clean_text(ws.cell(row=2, column=2).value),
        },
    }


def load_existing_payload() -> dict[str, Any]:
    if not OUTPUT.exists():
        return {}
    try:
        return json.loads(OUTPUT.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def main() -> None:
    wb = load_workbook(SOURCE, data_only=False)
    existing_payload = load_existing_payload()
    existing_modules = existing_payload.get("modules", {}) if isinstance(existing_payload, dict) else {}
    existing_handover = existing_modules.get("handover", {}) if isinstance(existing_modules, dict) else {}
    shipping_lines = []
    for ws in wb.worksheets:
        if ws.title == "ALL NAV":
            continue
        shipping_lines.append(extract_lines(ws, ws.title))

    module_defaults = {
        "settings": {
            "defaultQuoteCurrency": existing_handover.get("settings", {}).get("defaultQuoteCurrency", "MXN"),
            "defaultPriceMode": existing_handover.get("settings", {}).get("defaultPriceMode", "aftertax"),
        },
        "taxRatePresets": existing_handover.get("taxRatePresets", [
            {"id": "vat-0", "label": "0%", "rate": 0},
            {"id": "vat-16", "label": "16%", "rate": 0.16},
        ]),
    }

    payload = {
        "generatedFrom": str(SOURCE),
        "exchangeRates": existing_payload.get("exchangeRates", {
            "provider": "Frankfurter",
            "docsUrl": "https://frankfurter.dev/v1/",
            "asOfDate": None,
            "lastCheckedAt": None,
            "lastError": None,
            "defaultQuoteCurrency": "MXN",
            "pairs": [],
        }),
        "modules": {
            "handover": {
                **module_defaults,
                "shippingLines": shipping_lines,
            },
            "customs": existing_modules.get("customs", {
                **module_defaults,
                "shippingLines": [],
            }),
            "inland": existing_modules.get("inland", {
                **module_defaults,
                "shippingLines": [],
            }),
        },
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
